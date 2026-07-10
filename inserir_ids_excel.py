"""
Cruza nomes de pacientes entre:
  - aca.xlsx  (aba Listagem_pacientes, coluna Paciente)
  - pacientes.csv (colunas id, nome, telefone)

Insere coluna "ID_Simples" no Excel com o ID correspondente.
Estrategia de match:
  1. Exato (apos normalizar caixa e acentos)
  2. Fallback: telefone (se disponiveis)
"""

import csv
import unicodedata
import re
import openpyxl
from pathlib import Path

EXCEL_FILE = Path(__file__).parent / "aca.xlsx"
CSV_FILE   = Path(__file__).parent / "pacientes.csv"
SHEET_NAME = "Listagem_pacientes"
COL_PACIENTE = 1   # coluna A
COL_CELULAR  = 5   # coluna E


def normalizar(s: str) -> str:
    """Remove acentos, coloca em lowercase, elimina espacos extras."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s

def normalizar_tel(t) -> str:
    """Deixa so digitos."""
    if not t:
        return ""
    return re.sub(r"\D", "", str(t))


# ── Carrega CSV ──────────────────────────────────────────────────────────────
print("Carregando pacientes.csv ...")
csv_por_nome = {}   # nome_norm -> id
csv_por_tel  = {}   # tel_norm  -> id

with open(CSV_FILE, encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        pid  = row["id"].strip()
        nome = normalizar(row["nome"])
        tel  = normalizar_tel(row["telefone"])

        csv_por_nome[nome] = pid
        if tel:
            csv_por_tel[tel] = pid

print(f"  {len(csv_por_nome)} pacientes no CSV")

# ── Abre Excel ───────────────────────────────────────────────────────────────
print(f"Abrindo {EXCEL_FILE} ...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

# Descobre proxima coluna livre
next_col = ws.max_column + 1

# Cabecalho da nova coluna
ws.cell(row=1, column=next_col, value="ID_Simples")

# ── Cruzamento ───────────────────────────────────────────────────────────────
matched_nome = 0
matched_tel  = 0
sem_match    = 0

for row_idx in range(2, ws.max_row + 1):
    nome_xlsx = ws.cell(row=row_idx, column=COL_PACIENTE).value
    tel_xlsx  = ws.cell(row=row_idx, column=COL_CELULAR).value

    nome_norm = normalizar(nome_xlsx)
    tel_norm  = normalizar_tel(tel_xlsx)

    pid = None

    # 1. Match por nome
    if nome_norm in csv_por_nome:
        pid = csv_por_nome[nome_norm]
        matched_nome += 1
    # 2. Fallback: match por telefone (ultimos 9 digitos)
    elif tel_norm and tel_norm[-9:] in {k[-9:]: v for k, v in csv_por_tel.items()}:
        lookup = {k[-9:]: v for k, v in csv_por_tel.items()}
        pid = lookup.get(tel_norm[-9:])
        if pid:
            matched_tel += 1
    
    if not pid:
        sem_match += 1

    ws.cell(row=row_idx, column=next_col, value=pid)

# ── Salva ────────────────────────────────────────────────────────────────────
wb.save(EXCEL_FILE)

total = ws.max_row - 1
print(f"\n=== Resultado ===")
print(f"  Total pacientes no Excel : {total}")
print(f"  Match por nome           : {matched_nome}")
print(f"  Match por telefone       : {matched_tel}")
print(f"  Sem match                : {sem_match}")
print(f"\n  Coluna 'ID_Simples' inserida na coluna {next_col} da aba '{SHEET_NAME}'")
print(f"  Arquivo salvo: {EXCEL_FILE}")
