"""
Servidor local ACA v2 — com config, token, enviados e Excel sync
Rode: py servidor.py | Acesse: http://localhost:8765
"""
import http.server, json, os, random, subprocess, sys, threading, time, urllib.request, urllib.parse, webbrowser
from pathlib import Path
from datetime import date, datetime
try:
    import pyautogui
    pyautogui.FAILSAFE = True
    PYGUI_OK = True
except ImportError:
    PYGUI_OK = False

CAMPS_FILE = Path(__file__).parent / 'campanhas.json'

BASE   = Path(__file__).parent
PORT   = 8765
CONFIG_FILE = BASE / "config.json"
EXCEL_FILE  = BASE / "aca.xlsx"
WA_QUEUE_FILE = BASE / "wa_queue.json"

USERS_FILE = BASE / "aca.xlsx"
ORC_FILE = BASE / "orcamentos_abertos.json"
EVOL_FILE = BASE / "evolucoes_resultado.json"
DASH_FILE = BASE / "dashboard.html"

def _mtime_iso(path: Path):
    try:
        if path.exists():
            return datetime.fromtimestamp(path.stat().st_mtime).isoformat()
    except Exception:
        return None
    return None

def _now_iso():
    return datetime.now().isoformat()

# ─── Config ──────────────────────────────────────────────────────────────────
def load_config():
    if CONFIG_FILE.exists():
        cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        cfg.setdefault("valor_limpeza", 200)
        cfg.setdefault("valor_consulta", 250)
        cfg.setdefault("dias_limpeza_kpi", 90)
        cfg.setdefault("dias_consulta_kpi", 60)
        cfg.setdefault("dias_exclusao_radar_limpeza", cfg.get("dias_limpeza_kpi", 90))
        cfg.setdefault("delay_fechamento_s", 4)
        cfg.setdefault("fechar_aba_apos_envio", True)
        return cfg
    return {"token": "", "mensagem_template": "", "wa_coords": {"x":0,"y":0},
            "intervalo_envio_s": 3, "enviados": [], "valor_limpeza": 200, "valor_consulta": 250,
            "dias_limpeza_kpi": 90, "dias_consulta_kpi": 60,
            "dias_exclusao_radar_limpeza": 90,
            "delay_fechamento_s": 4, "fechar_aba_apos_envio": True}

def load_camps():
    if CAMPS_FILE.exists():
        return json.loads(CAMPS_FILE.read_text(encoding="utf-8"))
    return []

def save_camps(camps):
    CAMPS_FILE.write_text(json.dumps(camps, ensure_ascii=False, indent=2), encoding="utf-8")

def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")

def masked_token(tok):
    if not tok or len(tok) < 8: return tok
    return tok[:4] + "·" * (len(tok) - 8) + tok[-4:]

_orc_det_cache = {}

_wa_lock = threading.Lock()
_wa_worker_thread = None
_wa_session = {
    "status": "disconnected",   # disconnected | qr_required | connected | degraded
    "updated_at": None,
    "note": "Sessão não iniciada",
}
_wa_queue = {
    "running": False,
    "paused": False,
    "last_error": "",
    "updated_at": None,
    "items": [],
}

def _wa_snapshot_unlocked():
    items = _wa_queue.get("items", [])
    def _count(st):
        return sum(1 for it in items if it.get("status") == st)
    return {
        "running": bool(_wa_queue.get("running")),
        "paused": bool(_wa_queue.get("paused")),
        "last_error": _wa_queue.get("last_error") or "",
        "updated_at": _wa_queue.get("updated_at"),
        "total": len(items),
        "pending": _count("pending"),
        "processing": _count("processing"),
        "sent": _count("sent"),
        "failed": _count("failed"),
    }

def _save_wa_state_unlocked():
    payload = {
        "session": _wa_session,
        "queue": {
            "running": bool(_wa_queue.get("running")),
            "paused": bool(_wa_queue.get("paused")),
            "last_error": _wa_queue.get("last_error") or "",
            "updated_at": _wa_queue.get("updated_at"),
            "items": _wa_queue.get("items", []),
        }
    }
    WA_QUEUE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

def _load_wa_state():
    if not WA_QUEUE_FILE.exists():
        return
    try:
        payload = json.loads(WA_QUEUE_FILE.read_text(encoding="utf-8"))
        with _wa_lock:
            s = payload.get("session") or {}
            q = payload.get("queue") or {}
            _wa_session["status"] = s.get("status") or _wa_session["status"]
            _wa_session["updated_at"] = s.get("updated_at")
            _wa_session["note"] = s.get("note") or _wa_session["note"]
            _wa_queue["running"] = False
            _wa_queue["paused"] = False
            _wa_queue["last_error"] = q.get("last_error") or ""
            _wa_queue["updated_at"] = q.get("updated_at")
            _wa_queue["items"] = q.get("items") if isinstance(q.get("items"), list) else []
            _save_wa_state_unlocked()
    except Exception:
        return

def _normalize_phone(raw):
    d = "".join(ch for ch in str(raw or "") if ch.isdigit())
    if not d:
        return ""
    if d.startswith("55"):
        return d
    if len(d) in (10, 11):
        return "55" + d
    return d

def _build_wpp_url(phone, text):
    num = _normalize_phone(phone)
    msg = urllib.parse.quote(str(text or ""))
    return f"https://web.whatsapp.com/send?phone={num}&text={msg}"

def _send_wpp_item(item, cfg):
    if not PYGUI_OK:
        raise RuntimeError("pyautogui não instalado")

    wa_cfg = cfg.get("wa_coords", {}) or {}
    x = int(wa_cfg.get("x", 0) or 0)
    y = int(wa_cfg.get("y", 0) or 0)
    if not x or not y:
        raise RuntimeError("Coordenadas do botão Enviar não configuradas")

    delay_open = float(cfg.get("wa_delay_open_s", 1.5) or 1.5)
    delay_before_send = float(cfg.get("wa_delay_before_send_s", 2.0) or 2.0)
    delay_after_send = float(cfg.get("wa_delay_after_send_s", 1.0) or 1.0)
    intervalo_base = float(cfg.get("intervalo_envio_s", 4) or 4)
    close_tab = bool(cfg.get("fechar_aba_apos_envio", True))

    url = str(item.get("url") or "").strip()
    if not url:
        phone = item.get("celular") or item.get("phone") or ""
        if not _normalize_phone(phone):
            raise RuntimeError("Celular inválido para envio")
        url = _build_wpp_url(phone, item.get("mensagem") or item.get("message") or "")

    opened = webbrowser.open(url, new=1, autoraise=True)
    if not opened:
        try:
            os.startfile(url)
            opened = True
        except Exception:
            opened = False
    if not opened:
        raise RuntimeError("Não foi possível abrir o navegador padrão")

    time.sleep(max(0.6, delay_open))
    time.sleep(max(0.8, delay_before_send))
    pyautogui.moveTo(x, y, duration=0.12)
    pyautogui.click(x, y)
    time.sleep(0.3)
    pyautogui.press("enter")
    time.sleep(max(0.4, delay_after_send))
    if close_tab:
        pyautogui.hotkey("ctrl", "w")

    jitter = random.uniform(-0.5, 0.8)
    time.sleep(max(0.4, intervalo_base + jitter))

def _wa_pick_next_unlocked():
    for it in _wa_queue.get("items", []):
        if it.get("status") == "pending":
            return it
    return None

def _wa_worker_loop():
    while True:
        with _wa_lock:
            if not _wa_queue.get("running"):
                break
            if _wa_queue.get("paused"):
                _wa_queue["updated_at"] = _now_iso()
                _save_wa_state_unlocked()
                item = None
            else:
                item = _wa_pick_next_unlocked()
                if not item:
                    _wa_queue["running"] = False
                    _wa_queue["updated_at"] = _now_iso()
                    _save_wa_state_unlocked()
                    break
                item["status"] = "processing"
                item["started_at"] = _now_iso()
                _wa_queue["updated_at"] = _now_iso()
                _save_wa_state_unlocked()

        if not item:
            time.sleep(0.4)
            continue

        try:
            cfg = load_config()
            with _wa_lock:
                sess = _wa_session.get("status")
            if sess != "connected":
                raise RuntimeError("Sessão WA não está conectada")

            _send_wpp_item(item, cfg)

            patient_id = str(item.get("patient_id") or item.get("id_paciente") or "").strip()
            if patient_id:
                cfg = load_config()
                if patient_id not in cfg.get("enviados", []):
                    cfg.setdefault("enviados", []).append(patient_id)
                    save_config(cfg)
                try:
                    marcar_enviado_excel(patient_id)
                except Exception:
                    pass

            with _wa_lock:
                item["status"] = "sent"
                item["finished_at"] = _now_iso()
                item["error"] = ""
                _wa_queue["updated_at"] = _now_iso()
                _wa_queue["last_error"] = ""
                _save_wa_state_unlocked()
        except Exception as e:
            with _wa_lock:
                item["status"] = "failed"
                item["finished_at"] = _now_iso()
                item["error"] = str(e)
                _wa_queue["updated_at"] = _now_iso()
                _wa_queue["last_error"] = str(e)
                _save_wa_state_unlocked()

def _wa_start_worker_if_needed():
    global _wa_worker_thread
    with _wa_lock:
        should_start = bool(_wa_queue.get("running"))
    if not should_start:
        return
    if _wa_worker_thread and _wa_worker_thread.is_alive():
        return
    _wa_worker_thread = threading.Thread(target=_wa_worker_loop, daemon=True)
    _wa_worker_thread.start()

def _wa_add_items(body):
    raw_items = body.get("items")
    if not isinstance(raw_items, list):
        raw_items = [body]

    added = 0
    with _wa_lock:
        for r in raw_items:
            if not isinstance(r, dict):
                continue
            item = {
                "id": str(r.get("id") or f"wa-{int(time.time()*1000)}-{random.randint(100,999)}"),
                "patient_id": str(r.get("patient_id") or r.get("id_paciente") or "").strip(),
                "nome": str(r.get("nome") or "").strip(),
                "celular": str(r.get("celular") or r.get("phone") or "").strip(),
                "mensagem": str(r.get("mensagem") or r.get("message") or "").strip(),
                "url": str(r.get("url") or "").strip(),
                "status": "pending",
                "created_at": _now_iso(),
                "started_at": None,
                "finished_at": None,
                "error": "",
            }
            if not item["url"] and not _normalize_phone(item["celular"]):
                continue
            _wa_queue["items"].append(item)
            added += 1
        _wa_queue["updated_at"] = _now_iso()
        _save_wa_state_unlocked()
    return added

def sd_api_get(url: str, token: str, timeout=20):
    req = urllib.request.Request(url, headers={
        "x-auth-token": token,
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0",
        "Origin": "https://app.simplesdental.com",
    })
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read())

def _parse_br_date(s: str):
    try:
        return datetime.strptime(str(s or ""), "%d/%m/%Y").date()
    except Exception:
        return None

def _parse_iso_date(s: str):
    try:
        txt = str(s or "")
        if txt.endswith("Z"):
            txt = txt[:-1] + "+00:00"
        return datetime.fromisoformat(txt).date()
    except Exception:
        return None

def _to_float(v, default=0.0):
    try:
        return float(v)
    except Exception:
        return default

def escolher_orcamento(content, data_ref_br, valor_ref):
    data_ref = _parse_br_date(data_ref_br)
    melhor = None
    melhor_key = None
    for o in content or []:
        status = str(o.get("status") or "")
        rank_status = 0 if status == "EM_ABERTO" else 1 if status == "APROVADO" else 2
        d = _parse_iso_date(o.get("data"))
        diff_data = abs((d - data_ref).days) if (d and data_ref) else 9999
        diff_val = abs(_to_float(o.get("valorTotal"), 0.0) - _to_float(valor_ref, 0.0))
        key = (rank_status, diff_data, diff_val)
        if melhor is None or key < melhor_key:
            melhor = o
            melhor_key = key
    return melhor

# ─── Estado de progresso ─────────────────────────────────────────────────────
_prog = {"rodando": False, "etapa": "", "pct": 0, "log": [], "erro": None}
_lock = threading.Lock()

def set_prog(etapa, pct, log=None):
    with _lock:
        _prog["etapa"] = etapa
        _prog["pct"]   = pct
        if log: _prog["log"].append(log)

def run_pipeline(script_name):
    with _lock:
        _prog.update({"rodando": True, "etapa": "Iniciando...", "pct": 0, "log": [], "erro": None})
    try:
        set_prog("Analisando orçamentos...", 5)
        pre = subprocess.run(
            [sys.executable, "-u", str(BASE/"analisar_orcamentos.py")],
            cwd=str(BASE), capture_output=True, text=True, encoding="utf-8", errors="replace"
        )
        if pre.returncode != 0:
            msg = (pre.stderr or pre.stdout or "Falha em analisar_orcamentos.py").strip()
            raise RuntimeError(msg)

        set_prog("Coletando evoluções da API...", 10)
        env = dict(os.environ)
        env["PYTHONUNBUFFERED"] = "1"
        proc = subprocess.Popen(
            [sys.executable, "-u", str(BASE/script_name)], cwd=str(BASE), env=env,
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True,
            encoding="utf-8", errors="replace")
        for line in proc.stdout:
            line = line.strip()
            if not line: continue
            with _lock: _prog["log"].append(line)
            try:
                if "/" in line and "]" in line and "[" in line:
                    num = int(line.split("[")[1].split("/")[0])
                    den = int(line.split("/")[1].split("]")[0])
                    set_prog(f"Coletando {num}/{den}...", min(90, 10 + int(num/den*75)))
            except: pass
        proc.wait()
        if proc.returncode != 0:
            raise RuntimeError(f"{script_name} retornou código {proc.returncode}")

        set_prog("Gerando dashboard...", 95)
        post = subprocess.run(
            [sys.executable, "-u", str(BASE/"gerar_dashboard.py")],
            cwd=str(BASE), capture_output=True, text=True, encoding="utf-8", errors="replace"
        )
        if post.returncode != 0:
            msg = (post.stderr or post.stdout or "Falha em gerar_dashboard.py").strip()
            raise RuntimeError(msg)

        set_prog("Concluído!", 100, "✅ Dashboard atualizado!")
        with _lock: _prog["rodando"] = False
    except Exception as e:
        with _lock:
            _prog.update({"rodando": False, "erro": str(e)})
            _prog["log"].append(f"❌ Erro: {e}")

def run_local_pipeline():
    with _lock:
        _prog.update({"rodando": True, "etapa": "Iniciando...", "pct": 0, "log": [], "erro": None})
    try:
        set_prog("Lendo base local de orçamentos...", 30)
        pre = subprocess.run(
            [sys.executable, "-u", str(BASE/"analisar_orcamentos.py")],
            cwd=str(BASE), capture_output=True, text=True, encoding="utf-8", errors="replace"
        )
        if pre.returncode != 0:
            msg = (pre.stderr or pre.stdout or "Falha em analisar_orcamentos.py").strip()
            raise RuntimeError(msg)

        set_prog("Gerando dashboard com base local...", 80)
        post = subprocess.run(
            [sys.executable, "-u", str(BASE/"gerar_dashboard.py")],
            cwd=str(BASE), capture_output=True, text=True, encoding="utf-8", errors="replace"
        )
        if post.returncode != 0:
            msg = (post.stderr or post.stdout or "Falha em gerar_dashboard.py").strip()
            raise RuntimeError(msg)

        set_prog("Concluído!", 100, "✅ Reprocessamento local concluído")
        with _lock:
            _prog["rodando"] = False
    except Exception as e:
        with _lock:
            _prog.update({"rodando": False, "erro": str(e)})
            _prog["log"].append(f"❌ Erro: {e}")

def run_script_step(script_name, etapa, pct_ini, pct_fim, ok_msg):
    set_prog(etapa, pct_ini)
    res = subprocess.run(
        [sys.executable, "-u", str(BASE/script_name)],
        cwd=str(BASE), capture_output=True, text=True, encoding="utf-8", errors="replace"
    )
    out = ((res.stdout or "") + "\n" + (res.stderr or "")).splitlines()
    for ln in out:
        ln = ln.strip()
        if not ln:
            continue
        with _lock:
            _prog["log"].append(ln)
    if res.returncode != 0:
        msg = (res.stderr or res.stdout or f"Falha em {script_name}").strip()
        raise RuntimeError(msg)
    set_prog(etapa, pct_fim, ok_msg)

def run_update_job(job):
    with _lock:
        _prog.update({"rodando": True, "etapa": "Iniciando...", "pct": 0, "log": [], "erro": None})
    try:
        if job == "users":
            run_script_step("extrair_pacientes.py", "Obtendo dados remotos de usuários...", 10, 55, "✅ Dados remotos de usuários obtidos")
            run_script_step("inserir_ids_excel.py", "Atualizando dados locais de usuários...", 60, 100, "✅ Usuários atualizados no local")
        elif job == "orcamentos":
            set_prog("Obtendo dados remotos de orçamentos...", 10)
            run_script_step("analisar_orcamentos.py", "Atualizando dados locais de orçamentos...", 20, 60, "✅ Orçamentos atualizados no local")
            run_script_step("atualizar_tratamentos_orc.py", "Sincronizando tratamentos dos orçamentos (remoto -> local)...", 60, 90, "✅ Tratamentos de orçamentos sincronizados no local")
            run_script_step("gerar_dashboard.py", "Atualizando dados locais do dashboard...", 92, 100, "✅ Dashboard atualizado")
        elif job == "evolucoes":
            run_script_step("coletar_evolucoes.py", "Obtendo dados remotos de evoluções...", 10, 80, "✅ Dados remotos de evoluções obtidos")
            run_script_step("gerar_dashboard.py", "Atualizando dados locais de evoluções...", 85, 100, "✅ Evoluções atualizadas no local")
        elif job == "all":
            run_script_step("extrair_pacientes.py", "Obtendo dados remotos de usuários...", 5, 20, "✅ Dados remotos de usuários obtidos")
            run_script_step("inserir_ids_excel.py", "Atualizando dados locais de usuários...", 20, 35, "✅ Usuários atualizados no local")
            set_prog("Obtendo dados remotos de orçamentos...", 38)
            run_script_step("analisar_orcamentos.py", "Atualizando dados locais de orçamentos...", 40, 55, "✅ Orçamentos atualizados no local")
            run_script_step("atualizar_tratamentos_orc.py", "Sincronizando tratamentos dos orçamentos (remoto -> local)...", 55, 75, "✅ Tratamentos de orçamentos sincronizados no local")
            run_script_step("coletar_evolucoes.py", "Obtendo dados remotos de evoluções...", 75, 90, "✅ Dados remotos de evoluções obtidos")
            run_script_step("gerar_dashboard.py", "Atualizando dados locais do dashboard...", 92, 100, "✅ Dashboard atualizado")
        else:
            raise RuntimeError(f"Job inválido: {job}")

        with _lock:
            _prog["rodando"] = False
            if _prog.get("pct", 0) < 100:
                _prog["pct"] = 100
            _prog["etapa"] = "Concluído!"
    except Exception as e:
        with _lock:
            _prog.update({"rodando": False, "erro": str(e)})
            _prog["log"].append(f"❌ Erro: {e}")

# ─── Excel: marcar enviado ────────────────────────────────────────────────────
def marcar_enviado_excel(patient_id: str):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Listagem_pacientes"]
        # Encontra ou cria coluna Enviado_WA
        header_row = [c.value for c in ws[1]]
        if "Enviado_WA" not in header_row:
            col_idx = ws.max_column + 1
            ws.cell(row=1, column=col_idx, value="Enviado_WA")
        else:
            col_idx = header_row.index("Enviado_WA") + 1
        # Encontra ID_Simples column
        id_col = None
        for i, h in enumerate(header_row, 1):
            if h == "ID_Simples": id_col = i
        if id_col is None: return False
        # Busca o paciente e marca
        today = date.today().strftime("%d/%m/%Y")
        for row in ws.iter_rows(min_row=2):
            if str(row[id_col-1].value or "").strip() == str(patient_id):
                ws.cell(row=row[0].row, column=col_idx, value=today)
                break
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Excel error: {e}")
        return False

# ─── HTTP Handler ─────────────────────────────────────────────────────────────
class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, *a): pass

    def send_json(self, obj, code=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, path, ctype="text/html"):
        try:
            data = Path(path).read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", ctype + "; charset=utf-8")
            self.send_header("Content-Length", len(data))
            self.end_headers()
            self.wfile.write(data)
        except FileNotFoundError:
            self.send_response(404); self.end_headers()

    def read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(length)) if length else {}

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        p = parsed.path
        qs = urllib.parse.parse_qs(parsed.query)
        if p in ("/", "/dashboard.html"):
            self.send_file(BASE/"dashboard.html")
        elif p == "/api/progresso":
            with _lock: self.send_json(dict(_prog))
        elif p == "/api/status":
            self.send_json({"online": True})
        elif p == "/api/automacao_status":
            cfg = load_config()
            wa = cfg.get("wa_coords", {}) or {}
            x = int(wa.get("x", 0) or 0)
            y = int(wa.get("y", 0) or 0)
            coords_ok = bool(x and y)
            py_ok = bool(PYGUI_OK)
            ready = py_ok and coords_ok
            if ready:
                msg = "Automação pronta para envio Web"
            elif not py_ok and not coords_ok:
                msg = "Instale pyautogui e configure as coordenadas do botão Enviar"
            elif not py_ok:
                msg = "pyautogui não instalado"
            else:
                msg = "Coordenadas do botão Enviar não configuradas"
            self.send_json({
                "ok": True,
                "ready": ready,
                "pyautogui": py_ok,
                "coords": {"x": x, "y": y, "ok": coords_ok},
                "msg": msg,
            })
        elif p == "/api/config":
            cfg = load_config()
            cfg["token_masked"] = masked_token(cfg.get("token",""))
            cfg.pop("token", None)
            self.send_json(cfg)
        elif p == "/api/pos_mouse":
            if not PYGUI_OK:
                self.send_json({"ok": False, "msg": "pyautogui não instalado"}); return
            pos = pyautogui.position()
            self.send_json({"ok": True, "x": pos.x, "y": pos.y})
        elif p == "/api/campanhas":
            self.send_json(load_camps())
        elif p == "/api/update_dates":
            self.send_json({
                "ok": True,
                "datas": {
                    "users": _mtime_iso(USERS_FILE),
                    "orcamentos": _mtime_iso(ORC_FILE),
                    "evolucoes": _mtime_iso(EVOL_FILE),
                    "local": _mtime_iso(DASH_FILE),
                }
            })
        elif p == "/api/wa/status":
            with _wa_lock:
                self.send_json({
                    "ok": True,
                    "session": dict(_wa_session),
                    "queue": _wa_snapshot_unlocked(),
                })
        elif p == "/api/wa/queue/status":
            with _wa_lock:
                items = list(_wa_queue.get("items", []))
                self.send_json({
                    "ok": True,
                    "queue": _wa_snapshot_unlocked(),
                    "items": items[-200:],
                })
        elif p == "/api/orcamento_tratamentos":
            pid = str((qs.get("pid") or [""])[0]).strip()
            data_ref = str((qs.get("data") or [""])[0]).strip()
            valor_ref = _to_float((qs.get("valor") or [0])[0], 0.0)
            if not pid:
                self.send_json({"ok": False, "msg": "Parâmetro pid ausente"}); return

            cache_key = f"{pid}|{data_ref}|{valor_ref:.2f}"
            if cache_key in _orc_det_cache:
                self.send_json(_orc_det_cache[cache_key]); return

            cfg = load_config()
            tok = str(cfg.get("token") or "").strip()
            if not tok:
                self.send_json({"ok": False, "msg": "Token não configurado"}); return

            try:
                url_orcs = f"https://api.simplesdental.com/pacientes/{pid}/orcamentos?pageNumber=1&pageSize=30"
                orcs = sd_api_get(url_orcs, tok)
                content = orcs.get("content") or []
                if not content:
                    resp = {"ok": True, "msg": "Sem orçamentos para o paciente", "tratamentos": [], "total": 0}
                    _orc_det_cache[cache_key] = resp
                    self.send_json(resp); return

                best = escolher_orcamento(content, data_ref, valor_ref)
                if not best:
                    resp = {"ok": True, "msg": "Orçamento não localizado", "tratamentos": [], "total": 0}
                    _orc_det_cache[cache_key] = resp
                    self.send_json(resp); return

                orc_id = int(best.get("id") or 0)
                url_proc = f"https://api.simplesdental.com/orcamentos/{orc_id}/procedimentos?idPaciente={pid}&pageNumber=1"
                proc = sd_api_get(url_proc, tok)
                content_proc = proc.get("content") or []
                tratamentos = []
                for item in content_proc:
                    pinfo = item.get("procedimento") or {}
                    nome = str(pinfo.get("nome") or pinfo.get("nomeTuss") or "").strip()
                    if not nome:
                        continue
                    tratamentos.append({
                        "nome": nome,
                        "valor": _to_float(item.get("valor"), 0.0),
                    })

                resp = {
                    "ok": True,
                    "orcamento_id": orc_id,
                    "descricao": best.get("descricao") or "",
                    "status": best.get("status") or "",
                    "data": best.get("data") or "",
                    "valor": _to_float(best.get("valorTotal"), 0.0),
                    "tratamentos": tratamentos,
                    "total": len(tratamentos),
                }
                _orc_det_cache[cache_key] = resp
                self.send_json(resp)
            except Exception as e:
                self.send_json({"ok": False, "msg": f"Falha ao buscar detalhes: {e}"})
        else:
            fp = BASE / p.lstrip("/")
            if fp.exists():
                ct = "text/css" if p.endswith(".css") else "application/javascript"
                self.send_file(fp, ct)
            else:
                self.send_response(404); self.end_headers()

    def do_POST(self):
        p = self.path.split("?")[0]

        if p == "/api/atualizar":
            with _lock:
                if _prog["rodando"]:
                    self.send_json({"ok": False, "msg": "Já rodando"}); return
            threading.Thread(target=run_pipeline, args=("coletar_evolucoes.py",), daemon=True).start()
            self.send_json({"ok": True})

        elif p == "/api/reprocessar_local":
            with _lock:
                if _prog["rodando"]:
                    self.send_json({"ok": False, "msg": "Já rodando"}); return
            threading.Thread(target=run_local_pipeline, daemon=True).start()
            self.send_json({"ok": True})

        elif p == "/api/analise_completa":
            with _lock:
                if _prog["rodando"]:
                    self.send_json({"ok": False, "msg": "Já rodando"}); return
            threading.Thread(target=run_pipeline, args=("coletar_todos.py",), daemon=True).start()
            self.send_json({"ok": True})

        elif p == "/api/update_users":
            with _lock:
                if _prog["rodando"]:
                    self.send_json({"ok": False, "msg": "Já rodando"}); return
            threading.Thread(target=run_update_job, args=("users",), daemon=True).start()
            self.send_json({"ok": True})

        elif p == "/api/update_orcamentos":
            with _lock:
                if _prog["rodando"]:
                    self.send_json({"ok": False, "msg": "Já rodando"}); return
            threading.Thread(target=run_update_job, args=("orcamentos",), daemon=True).start()
            self.send_json({"ok": True})

        elif p == "/api/update_evolucoes":
            with _lock:
                if _prog["rodando"]:
                    self.send_json({"ok": False, "msg": "Já rodando"}); return
            threading.Thread(target=run_update_job, args=("evolucoes",), daemon=True).start()
            self.send_json({"ok": True})

        elif p == "/api/update_all":
            with _lock:
                if _prog["rodando"]:
                    self.send_json({"ok": False, "msg": "Já rodando"}); return
            threading.Thread(target=run_update_job, args=("all",), daemon=True).start()
            self.send_json({"ok": True})

        elif p == "/api/config":
            body = self.read_body()
            cfg  = load_config()
            if "token" in body and body["token"]: cfg["token"] = body["token"]
            if "mensagem_template" in body: cfg["mensagem_template"] = body["mensagem_template"]
            if "wa_coords" in body: cfg["wa_coords"] = body["wa_coords"]
            if "intervalo_envio_s" in body: cfg["intervalo_envio_s"] = body["intervalo_envio_s"]
            if "delay_fechamento_s" in body: cfg["delay_fechamento_s"] = body["delay_fechamento_s"]
            if "fechar_aba_apos_envio" in body: cfg["fechar_aba_apos_envio"] = body["fechar_aba_apos_envio"]
            if "valor_limpeza" in body: cfg["valor_limpeza"] = body["valor_limpeza"]
            if "valor_consulta" in body: cfg["valor_consulta"] = body["valor_consulta"]
            if "dias_limpeza_kpi" in body: cfg["dias_limpeza_kpi"] = body["dias_limpeza_kpi"]
            if "dias_consulta_kpi" in body: cfg["dias_consulta_kpi"] = body["dias_consulta_kpi"]
            if "dias_exclusao_radar_limpeza" in body: cfg["dias_exclusao_radar_limpeza"] = body["dias_exclusao_radar_limpeza"]
            save_config(cfg)
            self.send_json({"ok": True, "token_masked": masked_token(cfg["token"])})

        elif p == "/api/testar_token":
            body = self.read_body()
            tok  = body.get("token") or load_config().get("token","")
            url  = "https://api.simplesdental.com/pacientes/37768255/evolucoes?pageSize=1&pageNumber=1"
            try:
                req = urllib.request.Request(url, headers={
                    "x-auth-token": tok, "Accept": "application/json",
                    "User-Agent": "Mozilla/5.0", "Origin": "https://app.simplesdental.com"})
                with urllib.request.urlopen(req, timeout=10) as r:
                    r.read()
                self.send_json({"ok": True, "msg": "Token válido ✅"})
            except Exception as e:
                self.send_json({"ok": False, "msg": f"Token inválido: {e}"})

        elif p == "/api/marcar_enviado":
            body = self.read_body()
            pid  = str(body.get("id","")).strip()
            if not pid:
                self.send_json({"ok": False, "msg": "ID ausente"}); return
            cfg = load_config()
            if pid not in cfg["enviados"]:
                cfg["enviados"].append(pid)
                save_config(cfg)
            ok_excel = marcar_enviado_excel(pid)
            self.send_json({"ok": True, "excel": ok_excel})

        elif p == "/api/desmarcar_enviado":
            body = self.read_body()
            pid  = str(body.get("id","")).strip()
            cfg  = load_config()
            cfg["enviados"] = [x for x in cfg["enviados"] if x != pid]
            save_config(cfg)
            self.send_json({"ok": True})

        elif p == "/api/enviar_wpp":
            body  = self.read_body()
            url   = body.get("url", "")
            delay = int(body.get("delay", 4))
            x     = int(body.get("x", 0))
            y     = int(body.get("y", 0))
            if not url:
                self.send_json({"ok": False, "msg": "URL ausente"}); return
            if not PYGUI_OK:
                self.send_json({"ok": False, "msg": "pyautogui não instalado"}); return
            try:
                webbrowser.open(url)
                time.sleep(max(2, delay))
                pyautogui.click(x, y)
                time.sleep(0.6)
                self.send_json({"ok": True})
            except Exception as e:
                self.send_json({"ok": False, "msg": str(e)})

        elif p == "/api/clicar_mouse":
            body  = self.read_body()
            delay = int(body.get("delay", 4))
            x     = int(body.get("x", 0))
            y     = int(body.get("y", 0))
            if not PYGUI_OK:
                self.send_json({"ok": False, "msg": "pyautogui não instalado"}); return
            try:
                time.sleep(max(1, delay))
                pyautogui.click(x, y)
                self.send_json({"ok": True})
            except Exception as e:
                self.send_json({"ok": False, "msg": str(e)})

        elif p == "/api/enviar_wpp_web":
            body = self.read_body()
            url  = body.get("url", "")
            delay_open        = float(body.get("delay_open", 2))
            delay_before_send = float(body.get("delay_before_send", 5))
            delay_before_close= float(body.get("delay_before_close", 2))
            delay_after_send  = float(body.get("delay_after_send", 2))
            close_tab         = bool(body.get("close_tab", True))

            cfg = load_config()
            wa_cfg = cfg.get("wa_coords", {}) or {}
            try:
                x_in = int(body.get("x", 0) or 0)
            except Exception:
                x_in = 0
            try:
                y_in = int(body.get("y", 0) or 0)
            except Exception:
                y_in = 0

            x_cfg = int(wa_cfg.get("x", 0) or 0)
            y_cfg = int(wa_cfg.get("y", 0) or 0)
            x = x_in if x_in else x_cfg
            y = y_in if y_in else y_cfg

            if not url:
                self.send_json({"ok": False, "msg": "URL ausente"}); return
            if not PYGUI_OK:
                self.send_json({"ok": False, "msg": "pyautogui não instalado"}); return
            if not x or not y:
                self.send_json({"ok": False, "msg": "Coordenadas do botão Enviar não configuradas (configuração atual inválida)"}); return

            try:
                opened = webbrowser.open(url, new=1, autoraise=True)
                if not opened:
                    try:
                        os.startfile(url)
                        opened = True
                    except Exception:
                        opened = False

                if not opened:
                    self.send_json({"ok": False, "msg": "Não foi possível abrir o navegador padrão"}); return

                time.sleep(max(1.0, delay_open))
                time.sleep(max(1.0, delay_before_send))

                # Tenta o clique no botão Enviar e, em seguida, Enter como fallback.
                pyautogui.moveTo(x, y, duration=0.15)
                pyautogui.click(x, y)
                time.sleep(0.5)
                pyautogui.press("enter")

                if close_tab:
                    time.sleep(max(0.5, delay_before_close))
                    pyautogui.hotkey("ctrl", "w")

                time.sleep(max(0.5, delay_after_send))
                self.send_json({"ok": True})
            except Exception as e:
                self.send_json({"ok": False, "msg": str(e)})

        elif p == "/api/wa/session":
            body = self.read_body()
            status = str(body.get("status") or "").strip().lower()
            note = str(body.get("note") or "").strip()
            allowed = {"disconnected", "qr_required", "connected", "degraded"}
            if status not in allowed:
                self.send_json({"ok": False, "msg": "Status inválido"}); return
            with _wa_lock:
                _wa_session["status"] = status
                _wa_session["updated_at"] = _now_iso()
                _wa_session["note"] = note or _wa_session.get("note") or ""
                _save_wa_state_unlocked()
            self.send_json({"ok": True, "session": dict(_wa_session)})

        elif p == "/api/wa/queue/add":
            body = self.read_body()
            added = _wa_add_items(body)
            auto_start = bool(body.get("auto_start", False))
            if auto_start and added > 0:
                with _wa_lock:
                    _wa_queue["running"] = True
                    _wa_queue["paused"] = False
                    _wa_queue["updated_at"] = _now_iso()
                    _save_wa_state_unlocked()
                _wa_start_worker_if_needed()
            with _wa_lock:
                self.send_json({"ok": True, "added": added, "queue": _wa_snapshot_unlocked()})

        elif p == "/api/wa/queue/control":
            body = self.read_body()
            action = str(body.get("action") or "").strip().lower()
            with _wa_lock:
                if action == "start":
                    _wa_queue["running"] = True
                    _wa_queue["paused"] = False
                elif action == "pause":
                    _wa_queue["paused"] = True
                elif action == "resume":
                    _wa_queue["running"] = True
                    _wa_queue["paused"] = False
                elif action == "stop":
                    _wa_queue["running"] = False
                    _wa_queue["paused"] = False
                elif action == "clear":
                    _wa_queue["items"] = [it for it in _wa_queue.get("items", []) if it.get("status") == "processing"]
                    _wa_queue["running"] = False
                    _wa_queue["paused"] = False
                elif action == "retry_failed":
                    for it in _wa_queue.get("items", []):
                        if it.get("status") == "failed":
                            it["status"] = "pending"
                            it["error"] = ""
                            it["started_at"] = None
                            it["finished_at"] = None
                    _wa_queue["running"] = True
                    _wa_queue["paused"] = False
                else:
                    self.send_json({"ok": False, "msg": "Ação inválida"}); return

                _wa_queue["updated_at"] = _now_iso()
                _save_wa_state_unlocked()

            if action in {"start", "resume", "retry_failed"}:
                _wa_start_worker_if_needed()

            with _wa_lock:
                self.send_json({"ok": True, "queue": _wa_snapshot_unlocked()})

        elif p == "/api/campanhas_limpar":
            body = self.read_body()
            apenas_concluidas = bool(body.get("apenas_concluidas", True))
            camps = load_camps()
            before = len(camps)
            if apenas_concluidas:
                camps = [c for c in camps if str(c.get("status", "")).lower() == "ativa"]
            else:
                camps = []
            save_camps(camps)
            self.send_json({"ok": True, "removidas": max(0, before - len(camps)), "restantes": len(camps)})

        elif p == "/api/campanhas_excluir":
            body = self.read_body()
            camp_id = str(body.get("id", "")).strip()
            if not camp_id:
                self.send_json({"ok": False, "msg": "ID da campanha ausente"}); return
            camps = load_camps()
            before = len(camps)
            camps = [c for c in camps if str(c.get("id", "")).strip() != camp_id]
            save_camps(camps)
            self.send_json({"ok": True, "removidas": max(0, before - len(camps)), "restantes": len(camps)})

        elif p == "/api/nova_campanha":
            body  = self.read_body()
            nome  = body.get("nome", "Campanha sem nome")
            cfg   = load_config()
            camps = load_camps()
            enviados_body = body.get("enviados")
            if isinstance(enviados_body, list):
                enviados = [str(x).strip() for x in enviados_body if str(x).strip()]
            else:
                enviados = [str(x).strip() for x in cfg.get("enviados", []) if str(x).strip()]

            total_body = body.get("total_enviados")
            if total_body is None:
                total_enviados = len(enviados)
            else:
                try:
                    total_enviados = int(total_body)
                except Exception:
                    total_enviados = len(enviados)

            camp = {
                "id":            f"camp-{datetime.now().strftime('%Y%m%d%H%M%S')}",
                "nome":          nome,
                "mensagem":      cfg.get("mensagem_template", ""),
                "data_inicio":   datetime.now().isoformat(),
                "data_fim":      datetime.now().isoformat(),
                "status":        "concluída",
                "total_enviados": total_enviados,
                "enviados":      enviados,
            }
            camps.insert(0, camp)
            save_camps(camps)
            cfg["enviados"] = []
            save_config(cfg)
            self.send_json({"ok": True, "campanha": camp})

        else:
            self.send_response(404); self.end_headers()

# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import webbrowser
    _load_wa_state()
    _wa_start_worker_if_needed()
    server = http.server.HTTPServer(("localhost", PORT), Handler)
    url = f"http://localhost:{PORT}"
    print(f"")
    print(f"  ========================================")
    print(f"  ACA - Central de Inteligencia v2.1")
    print(f"  Servidor rodando em {url}")
    print(f"  ========================================")
    print(f"")
    threading.Timer(1.5, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Servidor encerrado.")
