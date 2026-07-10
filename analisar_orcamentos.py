"""
Analisa orcamentos em aberto e gera orcamentos_abertos.json
Inclui: mensal, reprovados, totais para o dashboard v2
"""
import openpyxl
from collections import defaultdict
import unicodedata, re, json
from pathlib import Path
from datetime import datetime

wb = openpyxl.load_workbook('aca.xlsx')
ws_pac = wb['Listagem_pacientes']
ws_orc = wb['Relatorio_orcamentos_por_status']

def norm(s):
    if not s: return ''
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s.lower().strip())

# Pacientes
pacientes = {}
for row in ws_pac.iter_rows(min_row=2, values_only=True):
    n = norm(row[0])
    cel = str(row[4] or '').strip()
    pid = str(row[5] or '').strip() if row[5] else ''
    pacientes[n] = {'celular': cel, 'id': pid, 'nome_orig': str(row[0] or '')}

total_pacientes = len(pacientes)

# Orçamentos
em_aberto = []
reprovados = 0
mensal = defaultdict(lambda: {'count': 0, 'valor': 0.0})

for row in ws_orc.iter_rows(min_row=2, values_only=True):
    data, nome, doc, cel, email, cel_resp, desc, status, valor = row
    st = str(status or '').strip()
    v  = float(valor or 0)

    # Parse data (DD/MM/YYYY ou outros formatos)
    mes_key = ''
    try:
        dt = datetime.strptime(str(data), '%d/%m/%Y')
        mes_key = dt.strftime('%Y-%m')
    except:
        try:
            dt = datetime.fromisoformat(str(data)[:10])
            mes_key = dt.strftime('%Y-%m')
        except:
            pass

    if st == 'Em aberto':
        nome_norm = norm(nome)
        pac = pacientes.get(nome_norm, {})
        em_aberto.append({
            'data': str(data or ''),
            'nome': str(nome or ''),
            'descricao': str(desc or '').strip(),
            'celular': str(cel or pac.get('celular', '') or ''),
            'valor': v,
            'id': pac.get('id', ''),
        })
        if mes_key:
            mensal[mes_key]['count'] += 1
            mensal[mes_key]['valor'] += v
    elif st == 'Reprovado':
        reprovados += 1

total_val = sum(r['valor'] for r in em_aberto)
com_cel   = sum(1 for r in em_aberto if r['celular'])

# Por faixa de valor (ordenadas logicamente)
faixas_ordem = [
    'Ate R$ 500',
    'R$ 500 a 1.000',
    'R$ 1.000 a 3.000',
    'R$ 3.000 a 5.000',
    'Acima de R$ 5.000'
]
faixas = {f: {'count': 0, 'valor': 0.0} for f in faixas_ordem}

for r in em_aberto:
    v = r['valor']
    if   v < 500:   f = 'Ate R$ 500'
    elif v < 1000:  f = 'R$ 500 a 1.000'
    elif v < 3000:  f = 'R$ 1.000 a 3.000'
    elif v < 5000:  f = 'R$ 3.000 a 5.000'
    else:           f = 'Acima de R$ 5.000'
    faixas[f]['count'] += 1
    faixas[f]['valor'] += v

# Série mensal ordenada (últimos 12 meses)
mensal_sorted = sorted(mensal.items())[-12:]
mensal_list = [{'mes': k, 'count': v['count'], 'valor': round(v['valor'], 2)}
               for k, v in mensal_sorted]

result = {
    'total': len(em_aberto),
    'valor_total': round(total_val, 2),
    'com_celular': com_cel,
    'reprovados': reprovados,
    'total_pacientes': total_pacientes,
    'faixas': [{'faixa': k, 'count': faixas[k]['count'], 'valor': round(faixas[k]['valor'], 2)}
               for k in faixas_ordem],
    'mensal': mensal_list,
    'lista': em_aberto
}

with open('orcamentos_abertos.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f'Em aberto: {len(em_aberto)} | Valor: R$ {total_val:,.2f} | Reprovados: {reprovados}')
print(f'Meses com dados: {len(mensal_list)} | Total pacientes: {total_pacientes}')
