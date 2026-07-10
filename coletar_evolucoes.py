"""
Coleta evolucoes e orcamentos da API Simples Dental
para os pacientes com orcamentos em aberto.

TOKENS: edite as variaveis abaixo com os valores do seu browser.
"""

import urllib.request
import urllib.error
import json
import time
import openpyxl
import unicodedata
import re
from pathlib import Path
from datetime import datetime, timezone

# ─── CONFIGURACAO ────────────────────────────────────────────────────────────
TOKEN = "QbDYIw4sWwebEvSjFEylCPHbCIDez9rIdxqx24vTt3Kv92VSA6rrDRStW6COH1Y4"

BASE          = Path(__file__).parent
EXCEL_FILE    = BASE / "aca.xlsx"
CACHE_FILE    = BASE / "evolucoes_cache.json"
ORC_PROC_FILE = BASE / "orcamentos_procedimentos.json"
CONFIG_FILE   = BASE / "config.json"

DELAY_ENTRE_REQUESTS = 0.3   # segundos entre chamadas (nao sobrecarregar API)
KEYWORDS_LIMPEZA     = ["limpeza", "profilaxia"]
# ─────────────────────────────────────────────────────────────────────────────

cfg = {}
if CONFIG_FILE.exists():
    try:
        cfg = json.loads(CONFIG_FILE.read_text(encoding='utf-8'))
    except Exception:
        cfg = {}

token_cfg = str(cfg.get('token') or '').strip()
if token_cfg:
    TOKEN = token_cfg

DIAS_EXCLUSAO_RADAR = int(cfg.get('dias_exclusao_radar_limpeza', cfg.get('dias_limpeza_kpi', 90)) or 90)

HEADERS = {
    "x-auth-token": TOKEN,
    "Accept": "application/json",
    "User-Agent": "Mozilla/5.0",
    "Origin": "https://app.simplesdental.com",
}

def api_get(url: str) -> dict | None:
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise
    except Exception as e:
        print(f"    ERRO {url}: {e}")
        return None

def html_to_text(html: str) -> str:
    text = re.sub(r'<[^>]+>', ' ', str(html))
    text = re.sub(r'&[a-z]+;', ' ', text)
    return re.sub(r'\s+', ' ', text).strip().lower()

def norm(s: str) -> str:
    if not s: return ''
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s.lower().strip())

def dias_atras(data_str: str) -> int:
    try:
        dt = datetime.fromisoformat(data_str.replace('Z', '+00:00'))
        return (datetime.now(timezone.utc) - dt).days
    except:
        return 9999

# ─── Leitura do Excel ────────────────────────────────────────────────────────
print("Lendo aca.xlsx ...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws_pac = wb['Listagem_pacientes']
ws_orc = wb['Relatorio_orcamentos_por_status']

# Monta dicionario de pacientes: nome_norm -> {id, celular, nome}
pacientes = {}
for row in ws_pac.iter_rows(min_row=2, values_only=True):
    nome = norm(row[0])
    pid  = str(row[5] or '').strip()
    cel  = str(row[4] or '').strip()
    if pid:
        pacientes[nome] = {'id': pid, 'celular': cel, 'nome_orig': str(row[0] or '')}

# Identifica IDs com orcamento em aberto
ids_em_aberto = {}
for row in ws_orc.iter_rows(min_row=2, values_only=True):
    data, nome, doc, cel, email, cel_resp, desc, status, valor = row
    if str(status or '').strip() != 'Em aberto':
        continue
    nome_norm = norm(nome)
    pac = pacientes.get(nome_norm, {})
    pid = pac.get('id', '')
    if pid and pid not in ids_em_aberto:
        ids_em_aberto[pid] = {
            'id': pid,
            'nome': str(nome or ''),
            'celular': str(cel or pac.get('celular', '') or ''),
            'valor_orcamento': float(valor or 0),
            'data_orcamento': str(data or ''),
        }

print(f"  {len(pacientes)} pacientes no sistema")
print(f"  {len(ids_em_aberto)} pacientes unicos com orcamento em aberto")

# ─── Carrega cache existente (para retomar se interrompido) ──────────────────
cache = {}
if CACHE_FILE.exists():
    cache = json.loads(CACHE_FILE.read_text(encoding='utf-8'))
    for pid, row in cache.items():
        if isinstance(row, dict) and not row.get('id'):
            row['id'] = str(pid)
    print(f"  Cache existente: {len(cache)} pacientes ja coletados")

# ─── Coleta de Evolucoes ─────────────────────────────────────────────────────
print("\nColetando evolucoes (somente pacientes com orcamento em aberto)...")
total = len(ids_em_aberto)
erros = 0

for idx, (pid, info) in enumerate(ids_em_aberto.items(), 1):
    if pid in cache:
        continue  # ja coletado

    url = f"https://api.simplesdental.com/pacientes/{pid}/evolucoes?pageSize=50&pageNumber=1&verHtml=true"
    data = api_get(url)

    if data is None:
        erros += 1
        cache[pid] = {**info, 'evolucoes': [], 'erro': True}
        continue

    evolucoes = data.get('content', [])

    # Analisa cada evolucao
    ultima_limpeza_data  = None
    ultima_limpeza_dias  = None
    ultima_evolucao_data = None
    ultima_evolucao_dias = None
    ultimo_proc          = ''

    for ev in evolucoes:
        ev_data = ev.get('data', '')
        ev_dias = dias_atras(ev_data)
        descricao_txt = html_to_text(ev.get('descricao', ''))

        if ultima_evolucao_data is None:
            ultima_evolucao_data = ev_data
            ultima_evolucao_dias = ev_dias
            ultimo_proc = descricao_txt[:120]

        if ultima_limpeza_data is None:
            if any(k in descricao_txt for k in KEYWORDS_LIMPEZA):
                ultima_limpeza_data = ev_data
                ultima_limpeza_dias = ev_dias

    # Score de prioridade
    score = 0
    score += 3   # tem orcamento em aberto (todos aqui tem)
    if info['valor_orcamento'] >= 1000: score += 1
    if info['valor_orcamento'] >= 5000: score += 1
    if ultima_limpeza_dias is not None and ultima_limpeza_dias > 90: score += 2
    if ultima_evolucao_dias is not None and ultima_evolucao_dias > 60: score += 1

    prioridade = 'ALTA' if score >= 5 else 'MEDIA' if score >= 3 else 'BAIXA'

    cache[pid] = {
        **info,
        'ultima_limpeza_data': ultima_limpeza_data,
        'ultima_limpeza_dias': ultima_limpeza_dias,
        'ultima_evolucao_data': ultima_evolucao_data,
        'ultima_evolucao_dias': ultima_evolucao_dias,
        'ultimo_proc': ultimo_proc,
        'total_evolucoes': len(evolucoes),
        'score': score,
        'prioridade': prioridade,
    }

    # Salva cache a cada 50 registros
    if idx % 50 == 0:
        CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding='utf-8')
        print(f"  [{idx}/{total}] {info['nome'][:40]} — prioridade: {prioridade}")

    time.sleep(DELAY_ENTRE_REQUESTS)

# Salva cache final
CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding='utf-8')

# ─── Gera JSON para o dashboard ──────────────────────────────────────────────
resultado = sorted(cache.values(), key=lambda x: -x.get('score', 0))
resultado_filtrado = []
excluidos_limpeza_recente = 0
for r in resultado:
    d_limpeza = r.get('ultima_limpeza_dias')
    if d_limpeza is not None and int(d_limpeza) <= DIAS_EXCLUSAO_RADAR:
        excluidos_limpeza_recente += 1
        continue
    resultado_filtrado.append(r)

print(f"\n=== Coleta concluida ===")
print(f"  Total processados : {len(cache)}")
print(f"  Erros             : {erros}")
print(f"  Excluidos (limpeza <= {DIAS_EXCLUSAO_RADAR}d): {excluidos_limpeza_recente}")
alta  = sum(1 for r in resultado_filtrado if r.get('prioridade') == 'ALTA')
media = sum(1 for r in resultado_filtrado if r.get('prioridade') == 'MEDIA')
baixa = sum(1 for r in resultado_filtrado if r.get('prioridade') == 'BAIXA')
print(f"  Alta prioridade   : {alta}")
print(f"  Media prioridade  : {media}")
print(f"  Baixa prioridade  : {baixa}")

# Salva para o dashboard
dashboard_data = {
    'atualizado': datetime.now().isoformat(),
    'total': len(resultado_filtrado),
    'range_exclusao_limpeza_dias': DIAS_EXCLUSAO_RADAR,
    'excluidos_limpeza_recente': excluidos_limpeza_recente,
    'alta': alta, 'media': media, 'baixa': baixa,
    'lista': resultado_filtrado
}
out = BASE / "evolucoes_resultado.json"
out.write_text(json.dumps(dashboard_data, ensure_ascii=False, indent=2), encoding='utf-8')
print(f"\n  Resultado salvo em: {out}")
print("  Execute gerar_dashboard.py para atualizar o dashboard.")
